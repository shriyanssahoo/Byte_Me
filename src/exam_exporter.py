"""
Export exam schedule and seating arrangements to Excel.
Format follows the PDF template provided.
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime


class ExamExporter:
    """Exports exam schedule and seating arrangements to Excel."""
    
    def __init__(self, schedule):
        self.schedule = schedule
    
    def export_all(self, output_dir='output/exams'):
        """Export all exam schedules and seating arrangements."""
        os.makedirs(output_dir, exist_ok=True)
        
        # Group seating arrangements by date
        arrangements_by_date = {}
        for arrangement in self.schedule.seating_arrangements:
            date_key = arrangement.date
            if date_key not in arrangements_by_date:
                arrangements_by_date[date_key] = []
            arrangements_by_date[date_key].append(arrangement)
        
        # Export one file per date
        for date, arrangements in arrangements_by_date.items():
            # Sort by session
            arrangements.sort(key=lambda a: (a.session, a.room.room_id))
            
            # Create filename
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            filename = f"{output_dir}/Exam_Seating_{date_obj.strftime('%d_%m_%Y')}.xlsx"
            
            self._export_date_seating(date, arrangements, filename)
            print(f"  ✓ Exported: {filename}")
        
        # Export student-wise exam schedule
        self._export_student_schedules(output_dir)
    
    def _export_date_seating(self, date, arrangements, filename):
        """Export seating arrangement for one date (all sessions and rooms)."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Group by session
        sessions = {}
        for arr in arrangements:
            if arr.session not in sessions:
                sessions[arr.session] = []
            sessions[arr.session].append(arr)
        
        # Create one sheet per room per session
        for session in ['FN', 'AN']:
            if session not in sessions:
                continue
            
            for arrangement in sessions[session]:
                sheet_name = f"{arrangement.room.room_id}_{session}"
                ws = wb.create_sheet(title=sheet_name)
                
                self._format_room_seating(ws, arrangement)
        
        wb.save(filename)
    
    def _format_room_seating(self, ws, arrangement):
        """Format one room's seating arrangement following PDF template."""
        room = arrangement.room
        date_obj = datetime.strptime(arrangement.date, '%Y-%m-%d')
        
        # Header: Date, Room, Session
        ws.merge_cells('A1:F1')
        cell = ws['A1']
        cell.value = f"Date {date_obj.strftime('%d/%m/%Y')}"
        cell.font = Font(size=12, bold=True)
        cell.alignment = Alignment(horizontal='left')
        
        ws.merge_cells('A2:F2')
        cell = ws['A2']
        cell.value = f"Room {room.room_id} session {arrangement.session}"
        cell.font = Font(size=12, bold=True)
        cell.alignment = Alignment(horizontal='left')
        
        current_row = 4
        
        # WINDOW at the front (before students)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws.cell(current_row, 1)
        cell.value = "WINDOW"
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        current_row += 1
        
        # Column labels
        ws.cell(current_row, 1, "COL1")
        ws.cell(current_row, 3, "COL2")
        ws.cell(current_row, 5, "COL3")
        for col in [1, 3, 5]:
            ws.cell(current_row, col).font = Font(bold=True)
            ws.cell(current_row, col).alignment = Alignment(horizontal='center')
        current_row += 1
        
        # All 8 rows of students
        for row in range(1, room.rows + 1):
            roll_numbers = []
            for col in range(1, room.columns + 1):
                left_student = arrangement.get_seat(row, col, 0)
                right_student = arrangement.get_seat(row, col, 1)
                
                left_roll = left_student.roll_number if left_student else ""
                right_roll = right_student.roll_number if right_student else ""
                
                roll_numbers.append(left_roll)
                roll_numbers.append(right_roll)
            
            # Write row
            for idx, roll in enumerate(roll_numbers):
                ws.cell(current_row, idx + 1, roll)
                ws.cell(current_row, idx + 1).alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        # Door label at the back (after all students)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws.cell(current_row, 1)
        cell.value = "DOOR"
        cell.font = Font(size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    def _export_student_schedules(self, output_dir):
        """Export student-wise exam schedules."""
        # Group students by section
        students_by_section = {}
        
        for exam in self.schedule.exams:
            for student in exam.students:
                section_key = f"{student.branch}-{student.section}"
                if section_key not in students_by_section:
                    students_by_section[section_key] = set()
                students_by_section[section_key].add(student)
        
        # Export one file per section
        for section_key, students in students_by_section.items():
            filename = f"{output_dir}/Student_Exam_Schedule_{section_key}.xlsx"
            self._export_section_schedule(section_key, list(students), filename)
            print(f"  ✓ Exported: {filename}")
    
    def _export_section_schedule(self, section_key, students, filename):
        """Export exam schedule for one section with one tab per subject."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sort students by roll number
        students.sort(key=lambda s: s.roll_number)
        
        # Get all exams for these students
        all_exams = set()
        for student in students:
            student_exams = self.schedule.get_student_exams(student)
            for exam in student_exams:
                all_exams.add(exam)
        
        # Sort exams by course code
        all_exams = sorted(all_exams, key=lambda e: e.course_code)
        
        if not all_exams:
            # No exams, create a default sheet
            ws = wb.create_sheet(title="No Exams")
            ws.cell(1, 1, "No exams scheduled for this section")
            wb.save(filename)
            return
        
        # Create one tab per subject/exam
        for exam in all_exams:
            # Sheet name: Course code (max 31 chars for Excel)
            sheet_name = exam.course_code[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Header with course info
            ws.merge_cells('A1:F1')
            cell = ws['A1']
            cell.value = f"{exam.course_code} - {exam.course_title}"
            cell.font = Font(size=14, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 25
            
            # Exam details
            date_obj = datetime.strptime(exam.date, '%Y-%m-%d')
            ws.merge_cells('A2:F2')
            cell = ws['A2']
            cell.value = f"Date: {date_obj.strftime('%d/%m/%Y')} | Session: {exam.session} | Time: {exam.start_time}-{exam.end_time}"
            cell.font = Font(size=11, bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[2].height = 20
            
            # Find room allocations for this exam
            rooms_for_exam = []
            for arrangement in self.schedule.seating_arrangements:
                if arrangement.date == exam.date and arrangement.session == exam.session:
                    # Check if any student from this section is in this room
                    for row in range(1, arrangement.room.rows + 1):
                        for col in range(1, arrangement.room.columns + 1):
                            for pos in [0, 1]:
                                student = arrangement.get_seat(row, col, pos)
                                if student and student in students:
                                    if arrangement.room.room_id not in rooms_for_exam:
                                        rooms_for_exam.append(arrangement.room.room_id)
                                    break
            
            ws.merge_cells('A3:F3')
            cell = ws['A3']
            cell.value = f"Rooms: {', '.join(rooms_for_exam) if rooms_for_exam else 'TBA'}"
            cell.font = Font(size=10, italic=True)
            cell.alignment = Alignment(horizontal="center")
            
            # Column headers for student list
            headers = ['Roll Number', 'Student Name', 'Room', 'Seat Info']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(5, col_idx, header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # List students taking this exam
            row = 6
            students_in_exam = [s for s in students if s in exam.students]
            students_in_exam.sort(key=lambda s: s.roll_number)
            
            for student in students_in_exam:
                # Find student's seat
                student_room = "TBA"
                seat_info = "TBA"
                
                for arrangement in self.schedule.seating_arrangements:
                    if arrangement.date == exam.date and arrangement.session == exam.session:
                        found = False
                        for r in range(1, arrangement.room.rows + 1):
                            for c in range(1, arrangement.room.columns + 1):
                                for p in [0, 1]:
                                    if arrangement.get_seat(r, c, p) == student:
                                        student_room = arrangement.room.room_id
                                        position = "Left" if p == 0 else "Right"
                                        seat_info = f"Row {r}, Col {c} ({position})"
                                        found = True
                                        break
                                if found:
                                    break
                            if found:
                                break
                
                ws.cell(row, 1, student.roll_number)
                ws.cell(row, 2, student.name)
                ws.cell(row, 3, student_room)
                ws.cell(row, 4, seat_info)
                
                # Add borders
                for col in range(1, 5):
                    ws.cell(row, col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")
                
                row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 25
        
        wb.save(filename)