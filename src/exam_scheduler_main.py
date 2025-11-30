import pandas as pd
import os
from datetime import datetime, timedelta
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- Configuration & Paths ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.join(SCRIPT_DIR, '..')
DATA_DIR = os.path.join(ROOT_DIR, 'data')
OUTPUT_DIR = os.path.join(ROOT_DIR, 'output', 'visual_exams')

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(DATA_DIR, exist_ok=True)

# --- Styles ---
BORDER_THIN = Side(border_style="thin", color="000000")
BORDER_MEDIUM = Side(border_style="medium", color="000000")
ALL_BORDER = Border(top=BORDER_THIN, left=BORDER_THIN, right=BORDER_THIN, bottom=BORDER_THIN)
OUTER_BORDER = Border(top=BORDER_MEDIUM, left=BORDER_MEDIUM, right=BORDER_MEDIUM, bottom=BORDER_MEDIUM)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_VERTICAL = Alignment(horizontal="center", vertical="center", textRotation=90)
FONT_BOLD = Font(bold=True)
FILL_GRAY = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# --- 1. DATA GENERATION ---
def generate_student_dataset():
    """Generates students based on: CSE=170, DSAI=85, ECE=85 per sem."""
    print("[1/4] Generating student dataset...")
    student_data = []
    
    # Logic: Sem 7 (22-series), Sem 5 (23-series), Sem 3 (24-series), Sem 1 (25-series)
    batches = [
        {'sem': 7, 'prefix': '22'},
        {'sem': 5, 'prefix': '23'},
        {'sem': 3, 'prefix': '24'},
        {'sem': 1, 'prefix': '25'}
    ]
    
    branches = [
        {'code': 'BCS', 'dept': 'CSE', 'count': 170},
        {'code': 'BDS', 'dept': 'DSAI', 'count': 85},
        {'code': 'BEC', 'dept': 'ECE', 'count': 85}
    ]
    
    for batch in batches:
        for branch in branches:
            for i in range(1, branch['count'] + 1):
                roll_no = f"{batch['prefix']}{branch['code']}{i:03d}"
                student_data.append({
                    'roll_number': roll_no,
                    'name': f"Student {roll_no}",
                    'branch': branch['dept'],
                    'section': 'A' if i <= 85 else 'B',
                    'semester': batch['sem']
                })
                
    df = pd.DataFrame(student_data)
    csv_path = os.path.join(DATA_DIR, 'generated_students.csv')
    df.to_csv(csv_path, index=False)
    return df

# --- 2. DATA LOADING ---
try:
    print("[2/4] Loading configuration files...")
    courses_df = pd.read_csv(os.path.join(DATA_DIR, 'course.csv'))
    rooms_df = pd.read_csv(os.path.join(DATA_DIR, 'exam_rooms.csv'))
    config_df = pd.read_csv(os.path.join(DATA_DIR, 'exam_config.csv'))
    config = dict(zip(config_df['parameter'], config_df['value']))
    students_df = generate_student_dataset()
    
    # Map (Dept, Sem) -> List of Roll Numbers
    students_map = students_df.groupby(['branch', 'semester'])['roll_number'].apply(list).to_dict()

except FileNotFoundError as e:
    print(f"FATAL ERROR: Missing file: {e}")
    exit(1)

# --- 3. EXAM SCHEDULING ---
def generate_schedule():
    print("[3/4] Generating Exam Schedule...")
    exam_schedule = []
    
    start_date = datetime.strptime(config['exam_start_date'], '%Y-%m-%d')
    current_date = start_date
    slot_cycle = ['Morning', 'Afternoon']
    
    valid_courses = courses_df[courses_df['Semester'].isin([1, 3, 5, 7])].copy()
    semesters = sorted(valid_courses['Semester'].unique())
    
    # Organize courses: {Sem: [List of Course Codes]}
    sem_course_map = {}
    for sem in semesters:
        sem_course_map[sem] = valid_courses[valid_courses['Semester'] == sem]['Course Code'].unique().tolist()
    
    max_courses = max(len(c) for c in sem_course_map.values()) if sem_course_map else 0
    
    for i in range(max_courses):
        for slot_name in slot_cycle:
            # Pick one course from each semester for this slot
            for sem in semesters:
                if sem_course_map[sem]:
                    course_code = sem_course_map[sem].pop(0)
                    course_details = valid_courses[valid_courses['Course Code'] == course_code].iloc[0]
                    
                    # Find all depts taking this course
                    involved = valid_courses[valid_courses['Course Code'] == course_code]
                    
                    for _, entry in involved.iterrows():
                        dept = entry['Department']
                        student_list = students_map.get((dept, sem), [])
                        if len(student_list) > 0:
                            exam_schedule.append({
                                'Date': current_date.strftime('%Y-%m-%d'),
                                'Slot': slot_name,
                                'Course_Code': course_code,
                                'Course_Name': course_details['Course Name'],
                                'Semester': sem,
                                'Department': dept,
                                'Student_Count': len(student_list)
                            })
            
            if slot_name == 'Afternoon':
                current_date += timedelta(days=1)
                
    return pd.DataFrame(exam_schedule)

# --- 4. VISUAL SEATING PLAN GENERATION ---
def draw_room_layout(writer, sheet_name, room_id, rows, cols, seating_data, meta):
    """
    Draws the specific visual layout with Window, Door, Board, and Double-Columns.
    seating_data: List of dicts {'left': roll, 'right': roll} per row/col block.
    """
    ws = writer.book.create_sheet(sheet_name)
    
    # --- A. Setup Metadata Header ---
    ws['C2'] = f"Room"
    ws['D2'] = room_id
    ws['D2'].border = ALL_BORDER
    ws['D2'].alignment = ALIGN_CENTER
    ws['D2'].font = FONT_BOLD
    
    # Board (Centered roughly)
    board_start_col = 3 + (cols // 2)
    ws.merge_cells(start_row=2, start_column=board_start_col, end_row=3, end_column=board_start_col+3)
    board_cell = ws.cell(row=2, column=board_start_col)
    board_cell.value = "BOARD"
    board_cell.alignment = ALIGN_CENTER
    board_cell.font = FONT_BOLD
    board_cell.border = OUTER_BORDER
    
    ws[f'{get_column_letter(cols*2 + 2)}2'] = "Date"
    ws[f'{get_column_letter(cols*2 + 3)}2'] = meta['date']
    ws[f'{get_column_letter(cols*2 + 2)}3'] = "Session"
    ws[f'{get_column_letter(cols*2 + 3)}3'] = meta['session']
    ws[f'{get_column_letter(cols*2 + 3)}3'].border = ALL_BORDER
    ws[f'{get_column_letter(cols*2 + 3)}3'].alignment = ALIGN_CENTER

    # --- B. Layout Constants ---
    START_ROW = 5
    WINDOW_COL = 1
    SEAT_START_COL = 3
    
    # --- C. Draw WINDOW (Left) ---
    ws.merge_cells(start_row=START_ROW, start_column=WINDOW_COL, end_row=START_ROW + rows, end_column=WINDOW_COL)
    win_cell = ws.cell(row=START_ROW, column=WINDOW_COL)
    win_cell.value = "WINDOW"
    win_cell.alignment = ALIGN_VERTICAL
    win_cell.border = ALL_BORDER
    
    # --- D. Draw Seating Grid ---
    current_excel_col = SEAT_START_COL
    
    for c_idx in range(cols): # Iterate Physical Columns (COL1, COL2...)
        # 1. Draw Header (Merged 2 cells)
        ws.merge_cells(start_row=START_ROW, start_column=current_excel_col, end_row=START_ROW, end_column=current_excel_col+1)
        header_cell = ws.cell(row=START_ROW, column=current_excel_col)
        header_cell.value = f"COL{c_idx + 1}"
        header_cell.alignment = ALIGN_CENTER
        header_cell.font = FONT_BOLD
        header_cell.border = ALL_BORDER
        
        # Apply border to the merged cell's partner
        ws.cell(row=START_ROW, column=current_excel_col+1).border = ALL_BORDER

        # 2. Fill Rows
        for r_idx in range(rows):
            current_excel_row = START_ROW + 1 + r_idx
            
            # Extract student data for this coordinate
            # We map (c_idx, r_idx) to our flat list or dict
            key = (c_idx, r_idx)
            pair = seating_data.get(key, {'left': '', 'right': ''})
            
            # Left Seat
            left_cell = ws.cell(row=current_excel_row, column=current_excel_col)
            left_cell.value = pair['left']
            left_cell.alignment = ALIGN_CENTER
            left_cell.border = ALL_BORDER
            
            # Right Seat (Gray Background)
            right_cell = ws.cell(row=current_excel_row, column=current_excel_col+1)
            right_cell.value = pair['right']
            right_cell.alignment = ALIGN_CENTER
            right_cell.border = ALL_BORDER
            right_cell.fill = FILL_GRAY
            
        current_excel_col += 2 # Move 2 Excel columns over for next Physical Column
        
    # --- E. Draw DOOR (Right) ---
    door_col = current_excel_col + 1
    ws.merge_cells(start_row=START_ROW, start_column=door_col, end_row=START_ROW + rows, end_column=door_col)
    door_cell = ws.cell(row=START_ROW, column=door_col)
    door_cell.value = "DOOR"
    door_cell.alignment = ALIGN_VERTICAL
    door_cell.border = ALL_BORDER

    # Autofit columns roughly
    for col in range(1, door_col + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12

def generate_seating_plans(schedule_df):
    print("[4/4] Generating Visual Seating Plans...")
    
    # Group by Date/Slot
    grouped = schedule_df.groupby(['Date', 'Slot'])
    rooms_list = rooms_df.to_dict('records')
    
    for (date, slot), group in grouped:
        print(f"      Processing {date} ({slot})...")
        
        # 1. Student Pooling Strategy
        # Identify "Largest Dept" (Right Seat) vs "Others" (Left Seat)
        dept_counts = group.groupby('Department')['Student_Count'].sum()
        if dept_counts.empty: continue
        largest_dept = dept_counts.idxmax()
        
        pool_A = [] # Largest Dept (e.g. CSE) -> Right Seat (Gray)
        pool_B = [] # Others (DSAI/ECE) -> Left Seat (White)
        
        for _, row in group.iterrows():
            rolls = students_map.get((row['Department'], row['Semester']), [])
            for roll in rolls:
                if row['Department'] == largest_dept:
                    pool_A.append(roll)
                else:
                    pool_B.append(roll)
        
        # 2. Allocate to Rooms
        filename = f"Seating_{date}_{slot}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        # Create Excel Writer using openpyxl directly via pandas
        writer = pd.ExcelWriter(filepath, engine='openpyxl')
        wb = writer.book
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
            
        current_room_idx = 0
        
        while (pool_A or pool_B) and current_room_idx < len(rooms_list):
            room_info = rooms_list[current_room_idx]
            room_id = room_info['room_id']
            # Physical rows and Physical columns (Blocks)
            p_rows = int(room_info['rows'])
            p_cols = int(room_info['columns'])
            
            seating_data = {} # Key: (col_idx, row_idx), Value: {left: roll, right: roll}
            
            # Fill Room Logic
            # We fill Column 1 (top to bottom), then Column 2...
            for c in range(p_cols):
                for r in range(p_rows):
                    left_student = ""
                    right_student = ""
                    
                    # Try to fill Left (Small depts) first, if empty, fill with A
                    if pool_B:
                        left_student = pool_B.pop(0)
                    elif pool_A:
                        left_student = pool_A.pop(0)
                        
                    # Try to fill Right (Large dept)
                    if pool_A:
                        right_student = pool_A.pop(0)
                    elif pool_B: # If A empty, use B
                        right_student = pool_B.pop(0)
                        
                    if left_student or right_student:
                        seating_data[(c, r)] = {'left': left_student, 'right': right_student}
            
            # Draw the visual sheet
            meta = {'date': date, 'session': slot}
            draw_room_layout(writer, room_id, room_id, p_rows, p_cols, seating_data, meta)
            
            current_room_idx += 1
            
        writer.close()
        
    print(f"Done! Files saved in: {OUTPUT_DIR}")

# --- EXECUTION ---
if __name__ == "__main__":
    schedule_df = generate_schedule()
    # Save raw schedule for reference
    schedule_df.to_csv(os.path.join(OUTPUT_DIR, 'Exam_Schedule_Master.csv'), index=False)
    
    # Generate the visual files
    generate_seating_plans(schedule_df)