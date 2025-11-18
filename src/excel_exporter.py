"""
src/excel_exporter.py
(Corrected to capitalize the session_type for display)
"""
# src/excel_exporter.py
import io
from typing import Union
# ... other imports like openpyxl
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import List, Dict, Set
from .models import Section, Classroom, Timetable, ScheduledClass
from . import utils
import random
import re

# --- Styling Constants ---
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
DAY_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
DAY_FONT = Font(bold=True, size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER_SIDE = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
BREAK_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
BREAK_FONT = Font(color="808080", size=9)


class ExcelExporter:
    def __init__(self, all_sections: List[Section], all_classrooms: List[Classroom], 
                 all_faculty_schedules: Dict[str, Timetable]):
        
        self.all_sections = all_sections
        self.all_classrooms = all_classrooms
        self.all_faculty_schedules = all_faculty_schedules
        self.course_color_map = self._generate_color_map()
        print("\nInitializing Excel Exporter...")

    def _generate_color_map(self) -> Dict[str, str]:
        course_codes: Set[str] = set()
        for section in self.all_sections:
            for day_schedule in section.timetable.grid:
                for slot in day_schedule:
                    if slot and slot.course.course_code not in ["LUNCH", "BREAK"]:
                        if slot.course.parent_pseudo_name:
                            course_codes.add(slot.course.parent_pseudo_name)
                        else:
                            course_codes.add(slot.course.course_code)
        
        color_map = {}
        for code in course_codes:
            r = random.randint(180, 240)
            g = random.randint(180, 240)
            b = random.randint(180, 240)
            color_map[code] = f"{r:02X}{g:02X}{b:02X}"
        
        color_map["LUNCH"] = "F0F0F0"
        color_map["BREAK"] = "FFFFFF"
        return color_map

    def _format_cell_content(self, s_class: ScheduledClass, view_type: str) -> str:
        if not s_class:
            return ""
        if s_class.course.course_code == "LUNCH":
            return "LUNCH"
        if s_class.course.course_code == "BREAK":
            return "BREAK"
            
        course_name = s_class.course.course_name
        room_str = ", ".join(s_class.room_ids)
        section_str = s_class.section_id
        
        instructor_str = ", ".join(s_class.instructors)
        # --- THIS IS THE FIX ---
        # Capitalize the lowercase session type for display
        session_str = f"({s_class.session_type.capitalize()})"
        # --- END OF FIX ---
        
        if s_class.course.parent_pseudo_name:
            course_name = s_class.course.parent_pseudo_name
            instructor_str = "" 
            if "elective" in course_name.lower():
                session_str = "(Elective)"
            else:
                session_str = "(Basket)"
                
        elif s_class.course.is_pseudo_basket:
            course_name = s_class.course.course_name
            instructor_str = ""
            room_str = "TBD"
            if "elective" in course_name.lower():
                session_str = "(Elective)"
            else:
                session_str = "(Basket)"

        if view_type == 'section':
            return f"{course_name}\n{session_str}\n{instructor_str}\n{room_str}".strip()
        elif view_type == 'faculty':
            return f"{course_name}\n{session_str}\n{section_str}\n{room_str}".strip()
        
        return f"{course_name}\n{section_str}"


    def _style_and_fill_sheet(self, ws: Worksheet, timetable: Timetable, view_type: str):
        ws.cell(row=1, column=1, value="Time / Day").fill = HEADER_FILL
        ws.cell(row=1, column=1).font = HEADER_FONT
        ws.column_dimensions['A'].width = 18
        
        time_slots = utils.get_time_slots_list()
        for c, time_str in enumerate(time_slots, start=2):
            cell = ws.cell(row=1, column=c, value=time_str)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)
            ws.column_dimensions[get_column_letter(c)].width = 8
            
        for r, day in enumerate(utils.DAYS, start=2):
            cell = ws.cell(row=r, column=1, value=day)
            cell.fill = DAY_FILL
            cell.font = DAY_FONT
            cell.alignment = CENTER_ALIGN
            ws.row_dimensions[r].height = 70
            
        for day_idx in range(len(utils.DAYS)):
            row_idx = day_idx + 2
            col_idx = 2
            
            while col_idx <= utils.TOTAL_SLOTS_PER_DAY + 1:
                slot_index = col_idx - 2
                s_class = timetable.grid[day_idx][slot_index]
                
                if not s_class:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = THIN_BORDER
                    col_idx += 1
                    continue
                
                duration = 1
                while (slot_index + duration < utils.TOTAL_SLOTS_PER_DAY and
                       timetable.grid[day_idx][slot_index + duration] == s_class):
                    duration += 1
                
                ws.merge_cells(
                    start_row=row_idx, start_column=col_idx,
                    end_row=row_idx, end_column=col_idx + duration - 1
                )
                
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = self._format_cell_content(s_class, view_type) # type: ignore
                cell.alignment = CENTER_ALIGN
                
                if s_class.course.course_code == "BREAK":
                    cell.fill = BREAK_FILL
                    cell.font = BREAK_FONT
                else:
                    color_key = s_class.course.course_code
                    if s_class.course.parent_pseudo_name:
                        color_key = s_class.course.parent_pseudo_name
                    elif s_class.course.is_pseudo_basket:
                        color_key = s_class.course.course_name
                    
                    color = self.course_color_map.get(color_key, "FFFFFF")
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                for c in range(col_idx, col_idx + duration):
                    ws.cell(row=row_idx, column=c).border = THIN_BORDER
                        
                col_idx += duration

    def export_department_timetables(self, filepath: Union[str, io.BytesIO]):
        print(f"Exporting department timetables to {filepath}...")
        wb = Workbook()
        
        if wb.active:
            wb.remove(wb.active)
        
        if not self.all_sections:
            ws = wb.create_sheet(title="No Sections Generated")
        else:
            for section in sorted(self.all_sections, key=lambda s: s.id):
                safe_title = re.sub(r'[\\/*?:\[\]]', '', section.id)[:31]
                ws = wb.create_sheet(title=safe_title)
                self._style_and_fill_sheet(ws, section.timetable, view_type='section')
            
        try:
            wb.save(filepath)
            print(f"Successfully saved department timetables to {filepath}")
        except PermissionError:
            print(f"FATAL ERROR: Could not save to {filepath}. Is the file open in Excel?")
        except Exception as e:
            print(f"FATAL ERROR: Could not save department timetables. {e}")

    def export_faculty_timetables(self, filepath: Union[str, io.BytesIO]):
        print(f"Exporting faculty timetables to {filepath}...")
        wb = Workbook()
        
        if not self.all_faculty_schedules:
            ws = wb.active
            if ws:
                ws.title = "No Faculty Scheduled"
            wb.save(filepath)
            return

        if "Sheet" in wb.sheetnames:
            default_sheet = wb["Sheet"]
            if default_sheet:
                wb.remove(default_sheet)
            
        for faculty_name, timetable in sorted(self.all_faculty_schedules.items()):
            safe_name = re.sub(r'[\\/*?:\[\]]', '', faculty_name)[:31]
            ws = wb.create_sheet(title=safe_name)
            self._style_and_fill_sheet(ws, timetable, view_type='faculty')
            
        try:
            wb.save(filepath)
            print(f"Successfully saved faculty timetables to {filepath}")
        except PermissionError:
            print(f"FATAL ERROR: Could not save to {filepath}. Is the file open in Excel?")
        except Exception as e:
            print(f"FATAL ERROR: Could not save faculty timetables. {e}")