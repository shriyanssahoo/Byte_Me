"""
Main entry point for the timetable scheduler.
"""

import sys
import os
import pandas as pd

# Add src to path so we can import our modules
sys.path.insert(0, os.path.dirname(__file__))

from data_loader import DataLoader
from scheduler import Scheduler


def generate_course_colors(courses):
    """Generate unique colors for each course."""
    colors = [
        'FFE6E6', 'E6F3FF', 'E6FFE6', 'FFF3E6', 'F3E6FF',
        'FFE6F3', 'E6FFFF', 'FFFFE6', 'FFE6CC', 'E6E6FF',
        'CCFFE6', 'FFE6E6', 'E6CCFF', 'FFCCCC', 'CCFFCC',
        'CCCCFF', 'FFCCFF', 'CCFFFF', 'FFFFCC', 'FFCCAA'
    ]
    
    course_color_map = {}
    unique_courses = list(set(course.code for course in courses))
    
    for i, course_code in enumerate(unique_courses):
        course_color_map[course_code] = colors[i % len(colors)]
    
    return course_color_map


def export_to_excel(timetable, courses, filename='output/timetable.xlsx'):
    """Export timetable to grid-style Excel file with all breaks shown."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    os.makedirs('output', exist_ok=True)
    
    # Generate color mapping
    course_colors = generate_course_colors(courses)
    
    # Get all unique student groups
    groups = set()
    for assignment in timetable.assignments:
        groups.add(assignment.student_group)
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define complete schedule with breaks
    complete_schedule = [
        ('09:00-10:00', 'class'),
        ('10:00-10:10', 'break'),
        ('10:10-11:10', 'class'),
        ('11:10-11:20', 'break'),
        ('11:20-12:20', 'class'),
        ('12:20-12:30', 'break'),
        ('12:30-13:20', 'class'),
        ('13:20-13:30', 'break'),
        ('13:30-14:00', 'lunch'),
        ('14:00-15:00', 'class'),
        ('15:00-15:10', 'break'),
        ('15:10-16:10', 'class'),
        ('16:10-16:20', 'break'),
        ('16:20-17:20', 'class'),
        ('17:20-17:30', 'break'),
        ('17:30-18:00', 'class'),
    ]
    
    # Create a sheet for each student group
    for group in sorted(groups):
        if '&' in group:  # Skip combined groups
            continue
        
        ws = wb.create_sheet(title=group)
        
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        # Create header row (Time slots with breaks)
        ws.cell(1, 1, "Day/Time").fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=11)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions['A'].width = 12
        
        for col_idx, (time_slot, slot_type) in enumerate(complete_schedule, start=2):
            cell = ws.cell(1, col_idx, time_slot)
            
            if slot_type == 'lunch':
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            elif slot_type == 'break':
                cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = 12
        
        # Fill in the grid
        for row_idx, day in enumerate(days, start=2):
            # Day name cell
            cell = ws.cell(row_idx, 1, day.upper())
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_idx].height = 60
            
            # Fill time slots
            for col_idx, (time_slot, slot_type) in enumerate(complete_schedule, start=2):
                cell = ws.cell(row_idx, col_idx)
                
                if slot_type == 'lunch':
                    cell.value = "LUNCH\nBREAK"
                    cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                elif slot_type == 'break':
                    cell.value = "Break"
                    cell.fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
                    cell.font = Font(size=8, italic=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                else:  # class slot
                    # Find assignment for this day and time
                    assignment = None
                    for a in timetable.assignments:
                        if (a.student_group == group and 
                            a.slot.day == day and 
                            f"{a.slot.start_time}-{a.slot.end_time}" == time_slot):
                            assignment = a
                            break
                    
                    if assignment:
                        # Format: Course Code\nCourse Title\nFaculty\nRoom
                        cell_text = f"{assignment.course.code}\n{assignment.course.title}\n{assignment.faculty.name}\n{assignment.room.room_id}"
                        cell.value = cell_text
                        
                        # Apply course color
                        color = course_colors.get(assignment.course.code, "FFFFFF")
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.font = Font(size=8, bold=False)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        # Empty slot
                        cell.value = ""
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                cell.border = thin_border
    
    # Create Faculty Sheets
    faculty_dict = {}
    for assignment in timetable.assignments:
        fid = assignment.faculty.faculty_id
        if fid not in faculty_dict:
            faculty_dict[fid] = {
                'name': assignment.faculty.name,
                'assignments': []
            }
        faculty_dict[fid]['assignments'].append(assignment)
    
    for faculty_id, faculty_data in sorted(faculty_dict.items()):
        # Sanitize sheet name
        sheet_name = faculty_data['name'][:28]
        if len(faculty_data['name']) > 28:
            sheet_name += "..."
        ws = wb.create_sheet(title=sheet_name)
        
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        # Header row
        ws.cell(1, 1, "Day/Time").fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.cell(1, 1).font = Font(color="FFFFFF", bold=True, size=11)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions['A'].width = 12
        
        for col_idx, (time_slot, slot_type) in enumerate(complete_schedule, start=2):
            cell = ws.cell(1, col_idx, time_slot)
            
            if slot_type == 'lunch':
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            elif slot_type == 'break':
                cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            
            cell.font = Font(color="FFFFFF", bold=True, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = 12
        
        # Fill grid
        for row_idx, day in enumerate(days, start=2):
            cell = ws.cell(row_idx, 1, day.upper())
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_idx].height = 60
            
            for col_idx, (time_slot, slot_type) in enumerate(complete_schedule, start=2):
                cell = ws.cell(row_idx, col_idx)
                
                if slot_type == 'lunch':
                    cell.value = "LUNCH\nBREAK"
                    cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                    cell.font = Font(bold=True, size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                elif slot_type == 'break':
                    cell.value = "Break"
                    cell.fill = PatternFill(start_color="D5DBDB", end_color="D5DBDB", fill_type="solid")
                    cell.font = Font(size=8, italic=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                else:  # class slot
                    # Find assignment
                    assignment = None
                    for a in faculty_data['assignments']:
                        if (a.slot.day == day and 
                            f"{a.slot.start_time}-{a.slot.end_time}" == time_slot):
                            assignment = a
                            break
                    
                    if assignment:
                        cell_text = f"{assignment.course.code}\n{assignment.course.title}\n{assignment.student_group}\n{assignment.room.room_id}"
                        cell.value = cell_text
                        
                        color = course_colors.get(assignment.course.code, "FFFFFF")
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.font = Font(size=8)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.value = ""
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                cell.border = thin_border
    
    wb.save(filename)
    print(f"\n✓ Grid-style timetable with breaks exported to {filename}")
    print(f"  → Created {len(wb.sheetnames)} sheets (student groups + faculty)")


def print_failed_schedules(failed_schedules):
    """Print courses that could not be scheduled."""
    if not failed_schedules:
        print("\n✓ All courses scheduled successfully!")
        return
    
    print(f"\n{'='*90}")
    print("⚠ COURSES THAT COULD NOT BE SCHEDULED".center(90))
    print('='*90)
    
    # Group by course
    failed_by_course = {}
    for item in failed_schedules:
        key = (item['course_code'], item['course_title'], item['student_group'])
        if key not in failed_by_course:
            failed_by_course[key] = []
        failed_by_course[key].append(item['lecture_num'])
    
    for (code, title, group), lectures in sorted(failed_by_course.items()):
        print(f"\n{code} - {title}")
        print(f"  Student Group: {group}")
        print(f"  Failed Lectures: {', '.join(map(str, lectures))}")
        print(f"  Total Failed: {len(lectures)}")


def print_timetable_by_group(timetable, student_group):
    """Print timetable for a specific student group."""
    assignments = timetable.get_assignments_by_group(student_group)
    
    if not assignments:
        print(f"\nNo classes scheduled for {student_group}")
        return
    
    print(f"\n{'='*90}")
    print(f"TIMETABLE FOR {student_group}".center(90))
    print('='*90)
    
    # Group by day
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    for day in days:
        day_assignments = [a for a in assignments if a.slot.day == day]
        
        if day_assignments:
            print(f"\n{day.upper()}")
            print('-' * 90)
            
            # Sort by time
            day_assignments.sort(key=lambda a: a.slot.start_time)
            
            lunch_shown = False
            
            for assignment in day_assignments:
                time = f"{assignment.slot.start_time}-{assignment.slot.end_time}"
                
                # Show lunch break once before afternoon classes
                if not lunch_shown and assignment.slot.start_time >= '14:00':
                    print(f"{'13:30-14:00':13} | {'LUNCH BREAK':^73} |")
                    lunch_shown = True
                
                print(f"{time} | "
                      f"{assignment.course.code:8} | {assignment.course.title:30} | "
                      f"{assignment.faculty.name:20} | {assignment.room.room_id}")


def print_timetable_by_faculty(timetable, faculty_id, faculty_name):
    """Print timetable for a specific faculty member."""
    assignments = timetable.get_assignments_by_faculty(faculty_id)
    
    if not assignments:
        print(f"\nNo classes scheduled for {faculty_name}")
        return
    
    print(f"\n{'='*90}")
    print(f"TIMETABLE FOR {faculty_name}".center(90))
    print('='*90)
    
    # Group by day
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    for day in days:
        day_assignments = [a for a in assignments if a.slot.day == day]
        
        if day_assignments:
            print(f"\n{day.upper()}")
            print('-' * 90)
            
            # Sort by time
            day_assignments.sort(key=lambda a: a.slot.start_time)
            
            for assignment in day_assignments:
                print(f"{assignment.slot.start_time}-{assignment.slot.end_time} | "
                      f"{assignment.course.code:8} | {assignment.course.title:30} | "
                      f"{assignment.student_group:15} | {assignment.room.room_id}")


def main():
    """Main function to run the scheduler."""
    print("=" * 90)
    print("AUTOMATED TIMETABLE SCHEDULER - IIIT Dharwad".center(90))
    print("Team: Byte Me".center(90))
    print("=" * 90)
    
    # Step 1: Load data
    print("\n📂 Loading data...")
    loader = DataLoader(data_dir='data')
    loader.load_all_data()
    
    if not loader.courses:
        print("\n❌ No courses found! Please create data/courses.csv")
        return
    
    # Step 2: Generate timetable
    scheduler = Scheduler(
        courses=loader.courses,
        faculty_dict=loader.faculty,
        rooms=loader.rooms,
        slots=loader.slots
    )
    
    timetable = scheduler.generate_timetable()
    
    # Step 3: Validate
    scheduler.validate_timetable()
    
    # Step 4: Show failed schedules
    failed_schedules = scheduler.get_failed_schedules()
    print_failed_schedules(failed_schedules)
    
    # Step 5: Display results for individual sections only
    # Get unique student groups (excluding combined A&B)
    groups = set()
    for course in loader.courses:
        if '&' in course.student_group:
            # Split "3rdSem-A&B" into "3rdSem-A" and "3rdSem-B"
            base = course.student_group.split('&')[0]
            groups.add(base)
            if base.endswith('-A'):
                groups.add(base[:-1] + 'B')
        else:
            groups.add(course.student_group)
    
    for group in sorted(groups):
        # Only show individual sections (A or B), not combined
        if '&' not in group:
            print_timetable_by_group(timetable, group)
    
    # Display faculty schedules
    print("\n\n" + "=" * 90)
    print("FACULTY SCHEDULES".center(90))
    print("=" * 90)
    
    for faculty_id, faculty in sorted(loader.faculty.items()):
        print_timetable_by_faculty(timetable, faculty_id, faculty.name)
    
    # Step 6: Export to Excel
    export_to_excel(timetable, loader.courses)
    
    print("\n" + "=" * 90)
    print("✓ TIMETABLE GENERATION COMPLETE!".center(90))
    print("=" * 90)


if __name__ == "__main__":
    main()